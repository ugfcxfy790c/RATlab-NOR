"""
Writes per-rat NOR results to an Excel workbook.

Layout: one sheet per rat. Within a sheet, one block per video/session
belonging to that rat, each block has two side-by-side tables (Object 1,
Object 2) listing every bout's start/stop/duration, followed by a summary
row with total bout count and total exploration time.

write_workbook() (used by main.py/batch_worker_process.py) always rebuilds
the whole file from the batch's in-memory results -- machine-generated
data only, no highlighting.

write_manual_review() (used by gui/manual_review_dialog.py) is the other
writer in this module: it edits a single video's block *in place* inside
an existing (or new) workbook, for a person re-scoring one video by hand
after spotting suspicious model output. It never touches other sessions'
blocks, and it never overwrites the original machine-generated numbers
outright -- the first time a video gets a manual review, its
machine-generated block is displaced sideways (starting at column 11)
into a red-highlighted "COMPUTED -- reference" copy, and the human-entered
bout data is written in its place (columns 1-8) with a pale yellow
highlight. Re-finalizing the same video again just replaces the human
block in place; the reference copy is left alone so it always reflects
the model's original output, not a previous manual edit.
"""

from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TITLE_FONT = Font(bold=True, size=13)
SESSION_FONT = Font(bold=True, size=11)
HEADER_FONT = Font(bold=True)
SUMMARY_FONT = Font(bold=True, italic=True)

# -- manual-review highlighting (write_manual_review / _parse_rat_sheet) --
# Pale yellow: a block a human has entered/confirmed by hand.
HUMAN_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
# Red: the original machine-generated block, displaced sideways for
# reference once a human review has overridden it in place.
COMPUTED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
# Column gap (in columns) between a session's primary block (1-8) and its
# displaced-computed reference block, if any.
REFERENCE_COL_OFFSET = 10


def _safe_sheet_name(name):
    bad = set('[]:*?/\\')
    cleaned = "".join(c for c in str(name) if c not in bad)
    return cleaned[:31] or "Rat"


def write_workbook(results, output_path, object_names=("Object 1", "Object 2")):
    """
    results: dict, rat_id -> list of session dicts, each:
        {
            "session_label": str,          # e.g. "359 novel a"
            "video_name": str,
            "bouts": {1: [Bout, ...], 2: [Bout, ...]},
        }
    """
    wb = Workbook()
    wb.remove(wb.active)

    for rat_id, sessions in sorted(results.items()):
        ws = wb.create_sheet(_safe_sheet_name(f"Rat {rat_id}"))
        ws.cell(row=1, column=1, value=f"Rat {rat_id}").font = TITLE_FONT

        row = 3
        for session in sessions:
            row = _write_session_block(ws, row, session, object_names)
            row += 2

        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 14

    if not wb.sheetnames:
        wb.create_sheet("No Data")

    wb.save(str(output_path))


def _write_session_block(ws, row, session, object_names):
    ws.cell(row=row, column=1, value=session.get("session_label", session.get("video_name", ""))).font = SESSION_FONT
    row += 1

    bouts = session["bouts"]
    obj1_bouts = bouts.get(1, [])
    obj2_bouts = bouts.get(2, [])

    header_row = row
    ws.cell(row=header_row, column=1, value=object_names[0]).font = HEADER_FONT
    ws.cell(row=header_row, column=5, value=object_names[1]).font = HEADER_FONT
    row += 1

    col_headers = ["Bout #", "Start (s)", "Stop (s)", "Duration (s)"]
    for i, h in enumerate(col_headers):
        c1 = ws.cell(row=row, column=1 + i, value=h)
        c1.font = HEADER_FONT
        c1.fill = HEADER_FILL
        c2 = ws.cell(row=row, column=5 + i, value=h)
        c2.font = HEADER_FONT
        c2.fill = HEADER_FILL
    row += 1

    n_rows = max(len(obj1_bouts), len(obj2_bouts), 1)
    for i in range(n_rows):
        if i < len(obj1_bouts):
            b = obj1_bouts[i]
            ws.cell(row=row, column=1, value=i + 1)
            ws.cell(row=row, column=2, value=round(b.start_s, 3))
            ws.cell(row=row, column=3, value=round(b.stop_s, 3))
            ws.cell(row=row, column=4, value=round(b.duration_s, 3))
        if i < len(obj2_bouts):
            b = obj2_bouts[i]
            ws.cell(row=row, column=5, value=i + 1)
            ws.cell(row=row, column=6, value=round(b.start_s, 3))
            ws.cell(row=row, column=7, value=round(b.stop_s, 3))
            ws.cell(row=row, column=8, value=round(b.duration_s, 3))
        row += 1

    total1 = sum(b.duration_s for b in obj1_bouts)
    total2 = sum(b.duration_s for b in obj2_bouts)

    ws.cell(row=row, column=1, value="Total bouts").font = SUMMARY_FONT
    ws.cell(row=row, column=2, value=len(obj1_bouts)).font = SUMMARY_FONT
    ws.cell(row=row, column=5, value="Total bouts").font = SUMMARY_FONT
    ws.cell(row=row, column=6, value=len(obj2_bouts)).font = SUMMARY_FONT
    row += 1
    ws.cell(row=row, column=1, value="Total time (s)").font = SUMMARY_FONT
    ws.cell(row=row, column=2, value=round(total1, 3)).font = SUMMARY_FONT
    ws.cell(row=row, column=5, value="Total time (s)").font = SUMMARY_FONT
    ws.cell(row=row, column=6, value=round(total2, 3)).font = SUMMARY_FONT
    row += 1

    return row


# ---------------------------------------------------------------------------
# Manual review: edit one video's block in an existing (or new) workbook,
# without disturbing any other session's data.
# ---------------------------------------------------------------------------

def _cell_is_filled(cell, fill):
    fg = getattr(cell.fill, "fgColor", None)
    target = getattr(fill.fgColor, "rgb", None)
    if fg is None or target is None:
        return False
    return str(fg.rgb) == str(target)


def _read_object_bouts(ws, header_row, col_start):
    """Read one object's bout table, starting just below `header_row`
    (the "Bout #" / "Start (s)" / ... row) in columns
    col_start..col_start+3. Stops at that column's "Total bouts" summary
    cell. Returns (bouts, total_bouts_row) -- bouts is a list of
    (start_s, stop_s) tuples; total_bouts_row lets the caller find the
    rest of the block ("Total time (s)" is always the next row down)."""
    bouts = []
    r = header_row + 1
    max_row = ws.max_row
    while r <= max_row + 1:
        c0 = ws.cell(row=r, column=col_start).value
        if c0 == "Total bouts":
            return bouts, r
        if c0 not in (None, ""):
            start_s = ws.cell(row=r, column=col_start + 1).value
            stop_s = ws.cell(row=r, column=col_start + 2).value
            if isinstance(start_s, (int, float)) and isinstance(stop_s, (int, float)):
                bouts.append((float(start_s), float(stop_s)))
        r += 1
    # Malformed/truncated block (no "Total bouts" row found) -- bail out
    # rather than looping past the sheet.
    return bouts, max_row


def _parse_rat_sheet(ws):
    """Read every session block currently in a rat's worksheet -- as
    written by write_workbook, the batch pipeline, or a previous
    write_manual_review call -- into an ordered dict: video stem/session
    label -> entry:
        {
            "object_names": (name1, name2),
            "current_bouts": {1: [(start_s, stop_s), ...], 2: [...]},
            "current_is_human": bool,
            "reference_bouts": {1: [...], 2: [...]} or None,
        }
    Order matches the order blocks appear top-to-bottom in the sheet.
    """
    sessions = OrderedDict()
    row = 3
    max_row = ws.max_row
    while row <= max_row:
        label = ws.cell(row=row, column=1).value
        if label in (None, ""):
            row += 1
            continue
        label = str(label)

        obj1_name = ws.cell(row=row + 1, column=1).value or "Object 1"
        obj2_name = ws.cell(row=row + 1, column=5).value or "Object 2"
        header_row = row + 2  # the "Bout #" / "Start (s)" / ... row

        obj1_bouts, summary_row = _read_object_bouts(ws, header_row, col_start=1)
        obj2_bouts, _ = _read_object_bouts(ws, header_row, col_start=5)
        block_end = summary_row + 1  # "Total time (s)" row

        is_human = _cell_is_filled(ws.cell(row=row, column=1), HUMAN_FILL)

        reference_bouts = None
        ref_col = 1 + REFERENCE_COL_OFFSET
        ref_label = ws.cell(row=row, column=ref_col).value
        if ref_label not in (None, ""):
            ref_obj1_name = ws.cell(row=row + 1, column=ref_col).value or obj1_name
            ref_obj2_name = ws.cell(row=row + 1, column=ref_col + 4).value or obj2_name
            ref_header_row = row + 2
            ref_obj1_bouts, ref_summary_row = _read_object_bouts(ws, ref_header_row, col_start=ref_col)
            ref_obj2_bouts, _ = _read_object_bouts(ws, ref_header_row, col_start=ref_col + 4)
            reference_bouts = {1: ref_obj1_bouts, 2: ref_obj2_bouts}
            block_end = max(block_end, ref_summary_row + 1)
            # ref_obj*_name are read but not stored -- the reference block
            # is always redrawn using the *current* object_names below, so
            # a rename between passes can't leave the two blocks
            # mismatched.
            del ref_obj1_name, ref_obj2_name

        sessions[label] = {
            "object_names": (obj1_name, obj2_name),
            "current_bouts": {1: obj1_bouts, 2: obj2_bouts},
            "current_is_human": is_human,
            "reference_bouts": reference_bouts,
        }
        row = block_end + 1

    return sessions


def _write_bout_block(ws, start_row, label, bouts, object_names, col_offset, fill, label_suffix=""):
    """Write one object-pair bout table (same layout as
    _write_session_block) starting at (start_row, 1 + col_offset).

    If `fill` is set, the *entire* 8-column-wide rectangle this block
    occupies (every row from the label down through the "Total time (s)"
    summary row, not just the specific cells that happen to get a value)
    is filled solid first -- a block whose two object tables have
    different bout counts otherwise ends up with a jagged, half-colored
    look (blank cells where the shorter column ran out of rows, or where
    a row/column simply has no content, like the label row's columns
    2-4). Filling the whole rectangle up front, then writing content on
    top, keeps the highlight a clean solid block regardless of how
    ragged the underlying data is. `fill=None` (an ordinary machine
    block) skips this and leaves cells unfilled/default as before.

    Returns the next free row below the block.
    """
    obj1_bouts = list(bouts.get(1, []))
    obj2_bouts = list(bouts.get(2, []))
    n_rows = max(len(obj1_bouts), len(obj2_bouts), 1)
    # label row + object-name header row + column-header row + n_rows of
    # bout data + 2 summary rows ("Total bouts", "Total time (s)").
    block_height = n_rows + 5

    if fill:
        for r in range(start_row, start_row + block_height):
            for c in range(1 + col_offset, 9 + col_offset):
                ws.cell(row=r, column=c).fill = fill

    row = start_row
    label_cell = ws.cell(row=row, column=1 + col_offset, value=f"{label}{label_suffix}")
    label_cell.font = SESSION_FONT
    row += 1

    header_row = row
    c1 = ws.cell(row=header_row, column=1 + col_offset, value=object_names[0])
    c1.font = HEADER_FONT
    c2 = ws.cell(row=header_row, column=5 + col_offset, value=object_names[1])
    c2.font = HEADER_FONT
    row += 1

    col_headers = ["Bout #", "Start (s)", "Stop (s)", "Duration (s)"]
    for i, h in enumerate(col_headers):
        h1 = ws.cell(row=row, column=1 + col_offset + i, value=h)
        h1.font = HEADER_FONT
        if not fill:
            h1.fill = HEADER_FILL
        h2 = ws.cell(row=row, column=5 + col_offset + i, value=h)
        h2.font = HEADER_FONT
        if not fill:
            h2.fill = HEADER_FILL
    row += 1

    for i in range(n_rows):
        if i < len(obj1_bouts):
            s, e = obj1_bouts[i]
            for j, v in enumerate([i + 1, round(s, 3), round(e, 3), round(e - s, 3)]):
                ws.cell(row=row, column=1 + col_offset + j, value=v)
        if i < len(obj2_bouts):
            s, e = obj2_bouts[i]
            for j, v in enumerate([i + 1, round(s, 3), round(e, 3), round(e - s, 3)]):
                ws.cell(row=row, column=5 + col_offset + j, value=v)
        row += 1

    total1 = sum(e - s for s, e in obj1_bouts)
    total2 = sum(e - s for s, e in obj2_bouts)

    for col, count in ((1 + col_offset, len(obj1_bouts)), (5 + col_offset, len(obj2_bouts))):
        ws.cell(row=row, column=col, value="Total bouts").font = SUMMARY_FONT
        ws.cell(row=row, column=col + 1, value=count).font = SUMMARY_FONT
    row += 1
    for col, total in ((1 + col_offset, round(total1, 3)), (5 + col_offset, round(total2, 3))):
        ws.cell(row=row, column=col, value="Total time (s)").font = SUMMARY_FONT
        ws.cell(row=row, column=col + 1, value=round(total, 3)).font = SUMMARY_FONT
    row += 1

    return row


def _write_session_pair(ws, row, label, entry):
    """Write one session's current block (columns 1-8) and, if present,
    its displaced computed-reference block (columns 11-18) side by side
    starting at the same row. Returns the next free row below both."""
    fill = HUMAN_FILL if entry["current_is_human"] else None
    cur_next = _write_bout_block(
        ws, row, label, entry["current_bouts"], entry["object_names"],
        col_offset=0, fill=fill,
    )
    ref_next = row
    if entry["reference_bouts"] is not None:
        ref_next = _write_bout_block(
            ws, row, label, entry["reference_bouts"], entry["object_names"],
            col_offset=REFERENCE_COL_OFFSET, fill=COMPUTED_FILL,
            label_suffix="  (COMPUTED -- reference, displaced by manual review)",
        )
    return max(cur_next, ref_next)


def write_manual_review(output_path, rat_id, session_label, bouts, object_names):
    """Write one video's hand-scored bout data into the same
    workbook/sheet the batch pipeline writes to, in-place -- see this
    module's docstring for the displace-computed-data-sideways /
    highlight-in-place behavior. Every other session already in the
    sheet (this rat's other videos) is preserved untouched.

    output_path: the workbook to edit -- same path the batch pipeline
        wrote to for this rat/job (e.g. "<group>_NOR_results.xlsx").
        Created fresh if it doesn't exist yet (e.g. a video was manually
        reviewed before the batch pipeline ever ran for it).
    rat_id: same value main.py/batch_worker_process.py's filename
        parsing would produce -- determines the sheet ("Rat <rat_id>").
    session_label: the video stem -- must match the label the batch
        pipeline used (Path(video_path).stem) for this to land on the
        same block instead of creating a new one.
    bouts: {1: [(start_s, stop_s), ...], 2: [...]} -- object index
        matches the position of that object's name in `object_names`
        (1-based, same convention as write_workbook/scoring.py).
    object_names: (name1, name2), e.g. ("novel", "original").
    """
    output_path = Path(output_path)
    wb, sessions, sheet_name = _open_rat_sheet_for_edit(output_path, rat_id)

    object_names = tuple(object_names)
    entry = sessions.get(session_label)
    if entry is None:
        entry = {
            "object_names": object_names,
            "current_bouts": {k: list(v) for k, v in bouts.items()},
            "current_is_human": True,
            "reference_bouts": None,
        }
    else:
        # Only capture a reference copy the *first* time a video is
        # manually reviewed -- if one's already there, it's the original
        # machine output and should never be replaced by a later human
        # edit (that would defeat its purpose as a fixed point of
        # comparison).
        if not entry["current_is_human"] and entry["reference_bouts"] is None:
            entry["reference_bouts"] = entry["current_bouts"]
        entry["object_names"] = object_names
        entry["current_bouts"] = {k: list(v) for k, v in bouts.items()}
        entry["current_is_human"] = True
    sessions[session_label] = entry

    _rewrite_rat_sheet(wb, sheet_name, rat_id, sessions)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


def write_computed_review(output_path, rat_id, session_label, bouts, object_names):
    """Replace one video's block with a freshly *computed* (non-human,
    unhighlighted) result -- same bouts/object_names shape as
    write_manual_review, but for correcting the machine output itself
    rather than recording a human re-score of it.

    Used by gui/batch_review_dialog.py's "Swap Novel/Original" action:
    after swapping which coordinate is novel vs. original for one video
    and re-scoring against the corrected object_coords (same .slp
    tracking data, no re-inference needed), this writes the corrected
    result back in place of the old (wrongly-labeled) one.

    Unlike write_manual_review, this unconditionally discards whatever
    was there before -- any prior human review AND any displaced
    computed reference -- rather than preserving/displacing it. Both are
    products of the old, wrong object labels, so neither is worth
    keeping once the labels are fixed; carrying either forward would
    just resurrect the same mistake under a different heading. Callers
    should confirm with the user first if there's something (a human
    review or a reference) this would discard -- see
    has_human_review()/has_computed_reference().
    """
    output_path = Path(output_path)
    wb, sessions, sheet_name = _open_rat_sheet_for_edit(output_path, rat_id)

    sessions[session_label] = {
        "object_names": tuple(object_names),
        "current_bouts": {k: list(v) for k, v in bouts.items()},
        "current_is_human": False,
        "reference_bouts": None,
    }

    _rewrite_rat_sheet(wb, sheet_name, rat_id, sessions)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


def revert_to_computed(output_path, rat_id, session_label):
    """Undo a manual review for one video: restores its original
    machine-computed block as the current (un-highlighted) block again,
    discarding the human-entered data and the red reference-displacement
    marker alongside it. The inverse of write_manual_review's first-time
    displace-sideways step.

    Raises ValueError if there's nothing to revert to -- either the
    workbook/sheet/video isn't there at all, or this video was never
    batch-processed (only ever manually entered from scratch), so there's
    no computed reference to fall back to. Callers (the GUI's "Revert to
    Computed Data" action) should check has_computed_reference() first to
    decide whether to offer this at all, and can still catch ValueError
    as a fallback.
    """
    output_path = Path(output_path)
    if not output_path.exists():
        raise ValueError(f"{output_path} does not exist -- nothing to revert.")

    sheet_name = _safe_sheet_name(f"Rat {rat_id}")
    wb = load_workbook(output_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No sheet found for Rat {rat_id} in {output_path.name}.")

    sessions = _parse_rat_sheet(wb[sheet_name])
    entry = sessions.get(session_label)
    if entry is None:
        raise ValueError(f"No data found for '{session_label}' in Rat {rat_id}'s sheet.")
    if entry["reference_bouts"] is None:
        raise ValueError(f"'{session_label}' has no computed reference data to revert to.")

    entry["current_bouts"] = entry["reference_bouts"]
    entry["current_is_human"] = False
    entry["reference_bouts"] = None
    sessions[session_label] = entry

    _rewrite_rat_sheet(wb, sheet_name, rat_id, sessions)
    wb.save(str(output_path))


def _open_rat_sheet_for_edit(output_path, rat_id):
    """Load (or create) the workbook, and parse whatever's currently in
    rat `rat_id`'s sheet (or start from an empty session list if the
    sheet/workbook doesn't exist yet). Returns (wb, sessions, sheet_name)
    -- callers mutate `sessions` and pass it to _rewrite_rat_sheet()."""
    if output_path.exists():
        wb = load_workbook(output_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    sheet_name = _safe_sheet_name(f"Rat {rat_id}")
    if sheet_name in wb.sheetnames:
        sessions = _parse_rat_sheet(wb[sheet_name])
    else:
        sessions = OrderedDict()
    return wb, sessions, sheet_name


def _rewrite_rat_sheet(wb, sheet_name, rat_id, sessions):
    """Replace rat `rat_id`'s sheet in `wb` with a fresh one built from
    `sessions` (see _parse_rat_sheet's docstring for its shape) -- shared
    by write_manual_review and revert_to_computed, both of which mutate a
    parsed session model then need to lay the whole sheet back out from
    scratch (see this module's docstring for why a full rebuild, not an
    in-place row patch, is what makes growing/shrinking blocks safe)."""
    if sheet_name in wb.sheetnames:
        index = wb.sheetnames.index(sheet_name)
        wb.remove(wb[sheet_name])
    else:
        index = len(wb.sheetnames)
    ws = wb.create_sheet(sheet_name, index)
    ws.cell(row=1, column=1, value=f"Rat {rat_id}").font = TITLE_FONT

    row = 3
    for label, sess in sessions.items():
        row = _write_session_pair(ws, row, label, sess)
        row += 2

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 14
    for col in range(1 + REFERENCE_COL_OFFSET, 9 + REFERENCE_COL_OFFSET):
        ws.column_dimensions[get_column_letter(col)].width = 14
    return ws


def has_human_review(output_path, rat_id, session_label):
    """True if `session_label` already has a human-reviewed (pale
    yellow) block in rat `rat_id`'s sheet -- i.e. Finalize-ing again
    would overwrite prior manual review data rather than displacing
    machine-generated data. False if the workbook, sheet, or session
    doesn't exist yet (nothing to overwrite)."""
    output_path = Path(output_path)
    if not output_path.exists():
        return False
    wb = load_workbook(output_path)
    sheet_name = _safe_sheet_name(f"Rat {rat_id}")
    if sheet_name not in wb.sheetnames:
        return False
    sessions = _parse_rat_sheet(wb[sheet_name])
    entry = sessions.get(session_label)
    return bool(entry and entry["current_is_human"])


def has_computed_reference(output_path, rat_id, session_label):
    """True if `session_label` has a displaced computed-reference block
    to revert to -- i.e. revert_to_computed() would succeed rather than
    raise ValueError."""
    output_path = Path(output_path)
    if not output_path.exists():
        return False
    wb = load_workbook(output_path)
    sheet_name = _safe_sheet_name(f"Rat {rat_id}")
    if sheet_name not in wb.sheetnames:
        return False
    sessions = _parse_rat_sheet(wb[sheet_name])
    entry = sessions.get(session_label)
    return bool(entry and entry["reference_bouts"] is not None)
